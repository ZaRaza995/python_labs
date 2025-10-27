import csv
from pathlib import Path
# Убедитесь, что openpyxl установлен: pip install openpyxl
import openpyxl
from openpyxl.utils import get_column_letter

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV-файл в XLSX-файл с автоподбором ширины колонок.
    :param csv_path: Путь к входному CSV-файлу.
    :param xlsx_path: Путь для сохранения выходного XLSX-файла.
    :raises FileNotFoundError: Если CSV-файл не найден.
    :raises ValueError: Если CSV-файл пуст.
    """
    csv_p = Path(csv_path)
    xlsx_p = Path(xlsx_path)
    # 1. Проверяем, что входной файл существует и является файлом
    if not csv_p.is_file():
        raise FileNotFoundError(f"Файл не найден или это папка: {csv_path}")
    # 2. Создаем новую Excel-книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 3. Читаем данные из CSV и добавляем их в лист
    with csv_p.open('r', encoding='utf-8', newline='') as f:
        reader = csv.reader(f)
        
        # Прочитаем первую строку, чтобы убедиться, что файл не пустой
        try:
            first_row = next(reader)
            ws.append(first_row)
        except StopIteration:
            raise ValueError(f"CSV-файл пуст: {csv_path}")
        # Добавляем остальные строки
        for row in reader:
            ws.append(row)

    # 4. Авто-подбор ширины колонок
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        # Проходим по всем ячейкам в колонке, чтобы найти максимальную длину
        for cell in ws[column_letter]:
            # Проверяем, есть ли в ячейке значение
            if cell.value:
                # Находим максимальную длину, сравнивая с текущей
                max_length = max(max_length, len(str(cell.value)))
        # Устанавливаем ширину: максимум из 8, или реальная длина + 2 для отступа
        adjusted_width = max(8, max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    # 5. Сохраняем XLSX-файл
    xlsx_p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_p))
    
if __name__ == "__main__":
    # Пример использования: создание XLSX из обновленного CSV
    csv_path = "src/data/lab05/out/people_from_json.csv"
    xlsx_path = "src/data/lab05/out/people_from_json.xlsx"
    
    try:
        print(f"Создание XLSX из {csv_path}")
        csv_to_xlsx(csv_path, xlsx_path)
        print(f"Успешно создан файл: {xlsx_path}")
    except Exception as e:
        print(f"Ошибка: {e}")
